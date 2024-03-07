import io
import os
import boto3
import pandas as pd
from openpyxl import load_workbook
import google.generativeai as genai
from dotenv import load_dotenv
import time

load_dotenv()

API_KEY = os.getenv("API_KEY")
AWS_INPUT_BUCKET = os.getenv("aws_input_bucket")
AWS_OUTPUT_BUCKET = os.getenv("aws_output_bucket")
AWS_ARCHIVE_BUCKET = os.getenv("aws_archive_bucket")
AWS_ACCESS_KEY_ID = os.getenv("aws_access_key_id")
AWS_SECRET_ACCESS_KEY = os.getenv("aws_secret_access_key")
# AWS_SESSION_TOKEN = os.getenv("AWS_SESSION_TOKEN")


genai.configure(api_key=API_KEY)

s3_client = boto3.client(
    "s3",
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY
    # aws_session_token=AWS_SESSION_TOKEN,
)


def upload_file_to_s3(username):
    # print("Uploading file to S3 bucket...")
    file = f"./static/uploads/{username}_input.xlsx"
    file_path = s3_client.upload_file(file, AWS_INPUT_BUCKET, f'{username}_input.xlsx')
    # print("File uploaded successfully to S3 bucket!")
    return file_path


def generate_bdd_scenario(username):
    s3_client_data = s3_client.get_object(Bucket=AWS_INPUT_BUCKET, Key=f'{username}_input.xlsx')
    contents = s3_client_data['Body'].read()  # your Excel's essence, pretty much a stream
    # Read in data_only mode to parse Excel after all formulae evaluated
    wb = load_workbook(filename=(io.BytesIO(contents)), data_only=True)
    model = genai.GenerativeModel("gemini-1.0-pro")
    sheet = wb.active
    responses = []
    for row in range(2, sheet.max_row + 1):
        # for row in range(1, sheet.max_row+1):
        prompt = sheet.cell(row, 1).value
        # Generate response
        convo = model.start_chat()
        convo.send_message("Generate BDD scenario in feature file format for the  user story " + prompt)
        response = convo.last.text
        # Save response
        responses.append(response)
        # print(prompt)
        # response = prompt
        # responses.append(response)
    df1 = pd.DataFrame(responses)
    with io.StringIO() as csv_buffer:
        df1.to_csv(csv_buffer, index=False)
        ts = str(int(round(time.time())))
        response = s3_client.put_object(
            Bucket=AWS_OUTPUT_BUCKET, Key=f"{username}_output_{ts}.csv", Body=csv_buffer.getvalue()
        )
        status = response.get("ResponseMetadata", {}).get("HTTPStatusCode")
        s3 = boto3.resource('s3')
        s3.Object(AWS_ARCHIVE_BUCKET, f'{username}_input_{ts}.xlsx').copy_from(
            CopySource=f'{AWS_INPUT_BUCKET}/{username}_input.xlsx')
        s3.Object(AWS_INPUT_BUCKET, f'{username}_input.xlsx').delete()
        url = f"https://{AWS_OUTPUT_BUCKET}.s3.amazonaws.com/{username}_output_{ts}.csv"
        if status == 200:
            # print(f"Successful S3 put_object response. Status - {status}")
            return url
        else:
            # print(f"Unsuccessful S3 put_object response. Status - {status}")
            return None
